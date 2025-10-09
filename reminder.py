import pandas as pd
import datetime
import smtplib
import logging
import re
import os
import imaplib
import email
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import parseaddr
from dotenv import load_dotenv
from openpyxl import load_workbook

# --- Load credentials securely ---
load_dotenv()
SENDER_EMAIL = os.getenv("EMAIL_ADDRESS")
SENDER_PASSWORD = os.getenv("EMAIL_PASSWORD")

# --- Setup logging ---
logging.basicConfig(
    filename="reminder.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# --- Email format validator ---
def is_valid_email(email_addr):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email_addr)

# --- Send email using SMTP ---
def send_email(to_email, subject, body):
    if not is_valid_email(to_email):
        logging.warning(f"Invalid email format: {to_email}")
        return False, "Invalid email format"

    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            response = server.sendmail(SENDER_EMAIL, to_email, msg.as_string())
            if response:
                logging.error(f"SMTP rejection for {to_email}: {response}")
                return False, str(response)
            else:
                logging.info(f"Email sent to {to_email}")
                return True, ""
    except Exception as e:
        logging.error(f"SMTP exception for {to_email}: {e}")
        return False, str(e)

# --- Check bounced emails via IMAP ---
def check_bounced_addresses():
    bounced = {}
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(SENDER_EMAIL, SENDER_PASSWORD)
        mail.select("inbox")

        status, data = mail.search(None, '(FROM "mailer-daemon@googlemail.com")')
        if status != "OK":
            logging.warning("IMAP search failed.")
            return bounced

        email_ids = data[0].split()[-10:]  # Check last 10 bounce messages
        for eid in email_ids:
            status, msg_data = mail.fetch(eid, "(RFC822)")
            if status != "OK":
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body += part.get_payload(decode=True).decode(errors="ignore")
            else:
                body = msg.get_payload(decode=True).decode(errors="ignore")

            # match = re.search(r"550.*(?:<([^>]+)>|to\s+([^\s]+))", body, re.IGNORECASE)
            # if match:
            #     bounced_email = match.group(1) or match.group(2)
            #     bounced_email = bounced_email.strip()
            # else:
            #     bounced_email = parseaddr(msg["To"])[1]
            # bounced[bounced_email] = "550 bounce detected"

            match = re.search(r"Your message wasn'?t delivered to ([^\s]+)", body)
            if match:
                bounced_email = match.group(1).strip()
                bounced[bounced_email] = "550 bounce detected"
            else:
                logging.warning("⚠️ Could not extract bounced email from message.")
                print("⚠️ Could not extract bounced email from message.")
            

    except Exception as e:
        logging.error(f"IMAP bounce check failed: {e}")

    # Debug printout
    print("📬 Bounced emails detected:")
    for email_addr, reason in bounced.items():
        print(f"- {email_addr}: {reason}")

    print(f"Bounced: ", bounced)
    return bounced

# --- Update Excel in-place ---
def update_events_excel(df, filename="events.xlsx"):
    try:
        book = load_workbook(filename)
        sheet = book.active

        for idx, row in df.iterrows():
            excel_row = idx + 2
            sheet.cell(row=excel_row, column=df.columns.get_loc("status") + 1).value = row["status"]
            sheet.cell(row=excel_row, column=df.columns.get_loc("last_checked") + 1).value = row["last_checked"]
            sheet.cell(row=excel_row, column=df.columns.get_loc("error") + 1).value = row["error"]

        book.save(filename)
        logging.info(f"✅ Updated '{filename}' with status, last_checked, and error info.")
        print(f"✅ Updated '{filename}' successfully.")
    except Exception as e:
        logging.error(f"❌ Failed to update '{filename}': {e}")
        print(f"❌ Failed to update '{filename}': {e}")

# --- Main reminder logic ---
def check_and_send_reminders():
    filename = "events.xlsx"
    required_columns = ["event_name", "event_date", "email", "status", "last_checked", "error"]

    if not os.path.exists(filename):
        logging.error(f"File '{filename}' not found.")
        print(f"❌ File '{filename}' not found.")
        return

    try:
        df = pd.read_excel(filename, parse_dates=["event_date"], dtype={"error": "string", "status": "string", "last_checked": "string"})
    except Exception as e:
        logging.error(f"Failed to read '{filename}': {e}")
        print(f"❌ Failed to read '{filename}': {e}")
        return

    if not all(col in df.columns for col in required_columns):
        logging.error(f"Missing required columns in '{filename}'. Expected: {required_columns}")
        print(f"❌ Missing required columns in '{filename}'. Expected: {required_columns}")
        return

    today = pd.Timestamp(datetime.date.today())
    timestamp_now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Fill missing values before processing
    df["last_checked"] = df["last_checked"].fillna("")
    df["status"] = df["status"].fillna("")
    df["error"] = df["error"].fillna("")

    for idx, row in df.iterrows():
        status = str(row.get("status")).strip().lower()
        if status in ["successful", "error"]:
            continue

        if pd.isnull(row["event_date"]) or not isinstance(row["event_date"], pd.Timestamp):
            df.at[idx, "status"] = "error"
            df.at[idx, "last_checked"] = timestamp_now
            df.at[idx, "error"] = "Invalid event date"
            logging.warning(f"Invalid event date skipped: {row}")
            continue

        days_until = (row["event_date"] - today).days
        df.at[idx, "days_until"] = days_until

        if days_until in [7, 2]:
            subject = f"Reminder: {row['event_name']} in {days_until} days"
            body = f"""Dear user,

This is a reminder that '{row['event_name']}' is scheduled for {row['event_date'].date()}.

Regards,
Your M.I.E.R.A.(MAY International Email Reminder Automation) Bot"""

            success, error_msg = send_email(row["email"], subject, body)
            df.at[idx, "status"] = "successful" if success else "error"
            df.at[idx, "last_checked"] = timestamp_now
            df.at[idx, "error"] = error_msg if not success else ""

    logging.info("⏳ Waiting 90 seconds for bounce messages...")
    time.sleep(90)

    bounced_emails = check_bounced_addresses()

    for idx, row in df.iterrows():
        if row["email"] in bounced_emails:
            df.at[idx, "status"] = "error"
            df.at[idx, "last_checked"] = timestamp_now
            df.at[idx, "error"] = bounced_emails[row["email"]]
            logging.warning(f"📩 Bounce override for {row['email']}")

    update_events_excel(df)

# --- Manual trigger ---
if __name__ == "__main__":
    check_and_send_reminders()